# -*- coding: utf-8 -*-
# =============================================================================
# --- EggScan 云端分析工具 (v3.1 - 符号清洗增强版) ---
# 【中文注释】增加了对AI输出中 Markdown 星号(*)的自动清洗，并保留期刊提取与样式优化。
# =============================================================================

import os
import re
import tempfile
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Flask, render_template, request, send_file, jsonify

# 全局变量声明
fitz = None
requests = None
pd = None
load_workbook = None
Font = None
Alignment = None
PatternFill = None
Border = None
Side = None

def import_heavy_libraries():
    """动态导入重量级库"""
    global fitz, requests, pd, load_workbook, Font, Alignment, PatternFill, Border, Side
    if fitz is None:
        print("正在加载核心分析库...")
        import fitz as _fitz
        import requests as _requests
        import pandas as _pd
        from openpyxl import load_workbook as _load_workbook
        from openpyxl.styles import Font as _Font, Alignment as _Alignment, PatternFill as _PatternFill, Border as _Border, Side as _Side
        
        fitz = _fitz
        requests = _requests
        pd = _pd
        load_workbook = _load_workbook
        Font = _Font
        Alignment = _Alignment
        PatternFill = _PatternFill
        Border = _Border
        Side = _Side
        print("✓ 分析库加载成功！")

# API和常量定义
LLM_URL = "https://api.deepseek.com/v1/chat/completions"

# 模式字段定义
SKIMMING_FIELDS = ["期刊", "研究问题", "核心论点", "研究方法", "关键结论", "相关性评估"]
INTENSIVE_FIELDS = ["期刊", "研究背景与缺口", "研究设计与方法", "主要结果与数据", "创新点与贡献", "局限性与批判", "可借鉴与启发"]

# 自定义模板
CUSTOM_TEMPLATE = """
请从以下角度分析这篇文献：
【研究主题】：文章的核心研究问题是什么？
【理论框架】：使用了什么理论基础？
【方法创新】：研究方法上有什么创新？
【数据质量】：数据来源和统计分析的可靠性如何？
【关键发现】：最重要的3个研究发现是什么？
【实践意义】：对实践有什么指导意义？

请用【字段名】：内容 的格式输出。
"""

# =============================================================================
# --- 核心辅助函数 ---
# =============================================================================

def smart_extract_text(pdf_path):
    """从PDF中智能提取文本"""
    try:
        doc = fitz.open(pdf_path)
        text = "\n".join([page.get_text() for page in doc])
        doc.close()
        text = re.sub(r'\n{3,}', '\n\n', text)
        text = re.sub(r' {2,}', ' ', text)
        return text
    except Exception as e:
        print(f"❌ 文本提取失败: {e}")
        return ""

def beautify_excel_professional(filepath):
    """专业的Excel美化"""
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        
        header_fill = PatternFill(fill_type="solid", fgColor="5B9BD5")
        header_font = Font(name='微软雅黑', bold=True, color="FFFFFF", size=11)
        
        # 【中文注释】正文字体设置为 微软雅黑 12号
        data_font = Font(name='微软雅黑', size=12)
        
        thin_border = Border(
            left=Side(style='thin', color='B4C6E7'),
            right=Side(style='thin', color='B4C6E7'),
            top=Side(style='thin', color='B4C6E7'),
            bottom=Side(style='thin', color='B4C6E7')
        )
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        ws.row_dimensions[1].height = 30
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        cell_value = str(cell.value)
                        chinese_chars = len(re.findall(r'[\u4e00-\u9fff]', cell_value))
                        other_chars = len(cell_value) - chinese_chars
                        effective_length = chinese_chars * 2 + other_chars
                        max_length = max(max_length, effective_length)
                except:
                    pass
            adjusted_width = min(max(max_length * 0.9, 15), 60)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            # 【中文注释】行高设置为 200磅
            ws.row_dimensions[row_num].height = 200
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                cell.border = thin_border
                cell.font = data_font
                if row_num % 2 == 0:
                    cell.fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
        
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        wb.save(filepath)
        print("✓ Excel美化完成")
        
    except Exception as e:
        print(f"⚠️ Excel美化失败: {e}")

# =============================================================================
# --- LLM调用函数 ---
# =============================================================================

def call_llm_for_mode(pdf_text, api_key, mode, language):
    """根据模式调用LLM"""
    
    if requests is None:
        import_heavy_libraries()
    
    lang_instruction = "Please output in English" if language == "English" else "请用中文输出"
    prompt_instruction = "请严格按照以下格式提取关键信息（每个字段必须填写，不要在答案中重复问题本身）："
    journal_info_prompt = "【期刊】：请从文献中提取发表的期刊名称和发表时间，格式要求：期刊名. 年份 月份（例如：Gastroenterology. 2021 October）。如果未找到则填“未知”。"
        
    if mode == '泛读模式':
        prompt = f"""你是一位专业的文献筛选专家，请对这篇论文进行快速泛读分析。{lang_instruction}\n{prompt_instruction}\n{journal_info_prompt}\n【研究问题】：这篇文章具体想回答什么问题？\n【核心论点】：作者最核心的观点是什么？\n【研究方法】：这是什么类型的研究？\n【关键结论】：最重要的研究结论是什么？\n【相关性评估】：评估其研究价值\n\n论文内容：\n{pdf_text[:25000]}"""
        fields = SKIMMING_FIELDS
    elif mode == '精读模式':
        prompt = f"""你是一位资深的学术研究专家，请对这篇论文进行深度精读分析。{lang_instruction}\n{prompt_instruction}\n{journal_info_prompt}\n【研究背景与缺口】：背景和空白\n【研究设计与方法】：样本、方法等\n【主要结果与数据】：关键数据\n【创新点与贡献】：理论或方法创新\n【局限性与批判】：作者承认的或你发现的问题\n【可借鉴与启发】：思路启发\n\n论文内容：\n{pdf_text[:35000]}"""
        fields = INTENSIVE_FIELDS
    elif mode == '自定义模式':
        prompt = f"""{journal_info_prompt}\n{CUSTOM_TEMPLATE}\n{lang_instruction}\n\n论文内容：\n{pdf_text[:30000]}"""
        fields = re.findall(r'【([^】]+)】', CUSTOM_TEMPLATE)
        if "期刊" not in fields: fields.insert(0, "期刊")
    else:
        return None, None
    
    headers = {"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"}
    payload = {
        "model": "deepseek-chat",
        "messages": [{"role": "system", "content": "你是专业的学术分析助手。不要使用Markdown加粗语法。"}, {"role": "user", "content": prompt}],
        "temperature": 0.1,
        "max_tokens": 3000
    }
    
    try:
        response = requests.post(LLM_URL, headers=headers, json=payload, timeout=280)
        response.raise_for_status()
        result = response.json()["choices"][0]["message"]["content"]
        # 清除常见的 Markdown 标题和分隔符
        result = re.sub(r'^\s*#+\s*|^\s*---\s*|\s*---\s*$', '', result, flags=re.MULTILINE)
        return result, fields
    except Exception as e:
        return f"API_ERROR: {e}", fields

def parse_llm_output(llm_text, fields):
    """解析LLM输出并进行内容清洗"""
    if llm_text.startswith("API_ERROR:"):
        return {field: llm_text if i == 0 else "API错误" for i, field in enumerate(fields)}
    
    # =====================================================================
    # ---【内容清理优化】---
    # 【中文注释】在此处删除所有的星号（*），防止 Markdown 加粗符号进入 Excel
    llm_text = llm_text.replace('*', '')
    # =====================================================================

    result_dict = {}
    chunks = re.split(r'(?=【.*?】)', llm_text)
    chunk_dict = {}
    for chunk in chunks:
        if not chunk.strip(): continue
        match = re.match(r'【(.*?)】[：:\s]*(.*)', chunk, re.DOTALL)
        if match:
            field_name, content = match.groups()
            content_lines = content.strip().split('\n')
            if len(content_lines) > 1 and ('什么' in content_lines[0] or '如何' in content_lines[0] or content_lines[0].endswith(('?', '？'))):
                content = '\n'.join(content_lines[1:]).strip()
            chunk_dict[field_name.strip()] = content.strip()
            
    for field in fields:
        result_dict[field] = chunk_dict.get(field, "未提取到")
    return result_dict

# =============================================================================
# --- Flask 应用主逻辑 ---
# =============================================================================

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/analyze', methods=['POST'])
def analyze_pdfs():
    import_heavy_libraries()
    pdf_files = request.files.getlist('pdfs')
    api_key = request.form.get('apiKey')
    mode = request.form.get('mode', '泛读模式')
    language = request.form.get('language', '中文')
    
    if not api_key: return jsonify({"error": "API密钥不能为空"}), 400
    
    all_results = []
    with ThreadPoolExecutor(max_workers=min(3, len(pdf_files))) as executor:
        futures = [executor.submit(process_single_pdf, f, api_key, mode, language) for f in pdf_files]
        for future in as_completed(futures):
            res = future.result()
            if res: all_results.append(res)
    
    if not all_results: return jsonify({"error": "处理失败"}), 500
    
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            df = pd.DataFrame(all_results)
            # 重新排列列顺序：文件名 -> 期刊 -> 其他内容 -> 分析时间
            if '文件名' in df.columns:
                cols = df.columns.tolist()
                cols.remove('文件名')
                cols.insert(0, '文件名')
                if '期刊' in df.columns:
                    cols.remove('期刊'); cols.insert(1, '期刊')
                if '分析时间' in df.columns:
                    cols.remove('分析时间'); cols.append('分析时间')
                df = df[cols]
            
            df.to_excel(tmp.name, index=False, engine='openpyxl')
            beautify_excel_professional(tmp.name)
            
            filename = f"EggScan_{mode}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            return send_file(tmp.name, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def process_single_pdf(pdf_file, api_key, mode, language):
    filename = pdf_file.filename
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        pdf_file.save(tmp.name)
        pdf_path = tmp.name
    try:
        text = smart_extract_text(pdf_path)
        if len(text.strip()) < 500: return None
        llm_output, fields = call_llm_for_mode(text, api_key, mode, language)
        result = parse_llm_output(llm_output, fields) if fields else {'分析结果': llm_output}
        result['文件名'] = filename
        result['分析时间'] = datetime.now().strftime("%Y-%m-%d %H:%M")
        return result
    except Exception as e:
        return {'文件名': filename, '错误': str(e)}
    finally:
        if os.path.exists(pdf_path): os.unlink(pdf_path)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
